"""Parse the XLSX spreadsheet and build the knowledge graph."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from aivc.graph import KnowledgeGraph
from aivc.models import (
    Company,
    Deal,
    PartnerEdge,
    Person,
    VCFirm,
)

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
XLSX_PATH = DATA_DIR / "raw" / "ai_vc_firms_worldwide.xlsx"
ALIASES_PATH = DATA_DIR / "aliases.json"

# Column indices for each sheet (0-based)
# Sheet: "AI VC Firms Directory"
FIRM_COLS = {
    "name": 0,
    "type": 1,
    "hq_region": 2,
    "hq_city": 3,
    "stage_focus": 4,
    "check_size": 5,
    "ai_focus": 6,
    "portfolio": 7,
    "aum": 8,
    "website": 9,
}

# Sheet: "Notable AI VC Investors"
INVESTOR_COLS = {
    "name": 0,
    "title": 1,
    "firm": 2,
    "linkedin": 3,
    "investments": 4,
    "notes": 5,
}

# Sheet: "AI VC News Sources"
NEWS_COLS = {
    "name": 0,
    "type": 1,
    "description": 2,
    "website": 3,
    "rss": 4,
    "newsletter": 5,
}


def load_aliases() -> dict[str, str]:
    """Load entity name aliases for normalization."""
    if ALIASES_PATH.exists():
        with open(ALIASES_PATH) as f:
            return json.load(f)
    return {}


def normalize_name(name: str, aliases: dict[str, str] | None = None) -> str:
    """Normalize an entity name using aliases and basic cleanup."""
    if not name:
        return name
    name = name.strip()
    if aliases:
        key = name.lower()
        if key in aliases:
            return aliases[key]
    return name


def parse_portfolio_list(raw: str) -> list[str]:
    """Parse comma-separated portfolio company names from a cell."""
    if not raw:
        return []
    # Split on commas, strip whitespace
    companies = [c.strip() for c in raw.split(",")]
    return [c for c in companies if c]


def ingest_xlsx(xlsx_path: str | Path | None = None) -> KnowledgeGraph:
    """Parse the full XLSX file and return a populated KnowledgeGraph."""
    xlsx_path = Path(xlsx_path) if xlsx_path else XLSX_PATH
    aliases = load_aliases()
    kg = KnowledgeGraph()
    source = f"xlsx:{xlsx_path.name}"

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    # ── Sheet 1: AI VC Firms Directory ──────────────────────────
    _ingest_firms(wb, kg, aliases, source)

    # ── Sheet 4: Notable AI VC Investors ────────────────────────
    _ingest_investors(wb, kg, aliases, source)

    # ── Sheet 5: AI VC News Sources (stored as metadata) ────────
    # News sources are not graph nodes but we could store them
    # For now, just parse for RSS feed URLs (used by scrapers)

    wb.close()
    return kg


def _ingest_firms(wb, kg: KnowledgeGraph, aliases: dict[str, str], source: str) -> None:
    """Ingest firms and their portfolio companies."""
    ws = wb["AI VC Firms Directory"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header

    for row in rows:
        if not row or not row[FIRM_COLS["name"]]:
            continue

        firm_name = str(row[FIRM_COLS["name"]]).strip()
        firm = VCFirm(
            name=firm_name,
            type=_cell(row, FIRM_COLS["type"]),
            hq_region=_cell(row, FIRM_COLS["hq_region"]),
            hq_city=_cell(row, FIRM_COLS["hq_city"]),
            stage_focus=_cell(row, FIRM_COLS["stage_focus"]),
            check_size=_cell(row, FIRM_COLS["check_size"]),
            ai_focus=_cell(row, FIRM_COLS["ai_focus"]),
            aum=_cell(row, FIRM_COLS["aum"]),
            website=_cell(row, FIRM_COLS["website"]),
            source=source,
        )
        kg.add_firm(firm)

        # Parse portfolio companies and add investment edges
        portfolio_raw = _cell(row, FIRM_COLS["portfolio"])
        companies = parse_portfolio_list(portfolio_raw)
        for company_name in companies:
            company_name = normalize_name(company_name, aliases)
            if company_name:
                kg.add_company(Company(name=company_name, source=source))
                kg.add_investment(firm_name, company_name, Deal(source=source))


def _ingest_investors(wb, kg: KnowledgeGraph, aliases: dict[str, str], source: str) -> None:
    """Ingest notable investors, their firm affiliations, and personal investments."""
    ws = wb["Notable AI VC Investors"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header

    for row in rows:
        if not row or not row[INVESTOR_COLS["name"]]:
            continue

        person_name = str(row[INVESTOR_COLS["name"]]).strip()
        title = _cell(row, INVESTOR_COLS["title"])
        firm_name = _cell(row, INVESTOR_COLS["firm"])
        linkedin = _cell(row, INVESTOR_COLS["linkedin"])
        investments_raw = _cell(row, INVESTOR_COLS["investments"])
        notes = _cell(row, INVESTOR_COLS["notes"])

        person = Person(
            name=person_name,
            title=title,
            linkedin=linkedin,
            notes=notes,
            source=source,
        )
        kg.add_person(person)

        # Link person to their firm
        if firm_name:
            firm_name = normalize_name(firm_name, aliases)
            kg.add_partner(
                person_name,
                firm_name,
                PartnerEdge(title=title, source=source),
            )

        # Parse personal notable investments
        if investments_raw:
            for company_name in parse_portfolio_list(investments_raw):
                company_name = normalize_name(company_name, aliases)
                if company_name:
                    from aivc.models import PersonalInvestmentEdge

                    kg.add_personal_investment(
                        person_name,
                        company_name,
                        PersonalInvestmentEdge(notes=notes, source=source),
                    )


def get_news_sources(xlsx_path: str | Path | None = None) -> list[dict]:
    """Parse news sources sheet and return list of source dicts (for scrapers)."""
    xlsx_path = Path(xlsx_path) if xlsx_path else XLSX_PATH
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["AI VC News Sources"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    sources = []
    for row in rows:
        if not row or not row[NEWS_COLS["name"]]:
            continue
        sources.append({
            "name": _cell(row, NEWS_COLS["name"]),
            "type": _cell(row, NEWS_COLS["type"]),
            "description": _cell(row, NEWS_COLS["description"]),
            "website": _cell(row, NEWS_COLS["website"]),
            "rss": _cell(row, NEWS_COLS["rss"]),
            "newsletter": _cell(row, NEWS_COLS["newsletter"]),
        })
    return sources


def ingest_deals_json(
    kg: KnowledgeGraph,
    deals_path: str | Path | None = None,
) -> int:
    """Ingest curated deals from a JSON file into the knowledge graph.

    Each deal record should have:
      - investor: str (VC firm name)
      - company: str (company name)
      - amount: str (e.g. "$500M")
      - round: str (e.g. "Series B")
      - date: str (e.g. "2025-06-15")
      - source_url: str (article URL where deal was reported)
      - company_sector: str (optional, e.g. "Foundation Models")

    Returns the number of deals ingested.
    """
    deals_path = Path(deals_path) if deals_path else DATA_DIR / "raw" / "curated_deals.json"
    if not deals_path.exists():
        return 0

    with open(deals_path) as f:
        deals_list = json.load(f)

    if not isinstance(deals_list, list):
        return 0

    aliases = load_aliases()
    count = 0

    for record in deals_list:
        investor = record.get("investor", "").strip()
        company = record.get("company", "").strip()
        if not investor or not company:
            continue

        investor = normalize_name(investor, aliases)
        company = normalize_name(company, aliases)

        source_url = record.get("source_url", "")
        source = f"curated:{source_url}" if source_url else "curated"

        # Add/merge the VC firm node
        kg.add_firm(VCFirm(name=investor, source=source))

        # Add/merge the company node (with optional sector)
        sector = record.get("company_sector", "")
        kg.add_company(Company(name=company, sector=sector, source=source))

        # Add the investment edge with deal details
        deal = Deal(
            amount=record.get("amount", ""),
            round=record.get("round", ""),
            date=record.get("date", ""),
            source=source,
        )
        kg.add_investment(investor, company, deal)
        count += 1

    return count


def _cell(row: tuple, idx: int) -> str:
    """Safely extract a cell value as a string."""
    if idx >= len(row) or row[idx] is None:
        return ""
    return str(row[idx]).strip()
